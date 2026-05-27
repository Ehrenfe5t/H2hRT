#!/usr/bin/env python3
"""
Convert measurement Excel to Precise simulation rx_list JSON file.

Coordinate transform: project_x=excel_x, project_y=excel_z, project_z=-excel_y

Usage:
  python tools/excel_to_rxlist.py "demo/412/Tx1 20dBm 3GHz .xlsx"
  python tools/excel_to_rxlist.py "demo/412/Tx1 20dBm 3GHz .xlsx" -o configs/app/rxlist_tx1.json
"""

import sys, re, json, argparse
from pathlib import Path
try:
    import openpyxl
except ImportError:
    print("pip install openpyxl"); sys.exit(1)


def extract_number(label):
    m = re.search(r"(\d+)", str(label))
    return int(m.group(1)) if m else 0


def main():
    parser = argparse.ArgumentParser(
        description="Convert measurement Excel to Precise simulation rx_list JSON")
    parser.add_argument("excel_path", nargs="?", default="G:/RT/H2hRT-7.1-SBR-/demo/412/Tx1 20dBm 3GHz .xlsx",
                        help="Excel file (default: demo/412/Tx1 20dBm 3GHz .xlsx)")
    parser.add_argument("--output", "-o", type=str, default="G:/RT/H2hRT-7.1-SBR-/demo/412/rx_tx1.json",
                        help="Output JSON (default: G:/RT/H2hRT-7.1-SBR-/configs/app/rx_tx1.json)")
    args = parser.parse_args()

    xlsx = Path(args.excel_path)
    if not xlsx.exists():
        print(f"ERROR: file not found: {xlsx}"); sys.exit(1)

    wb = openpyxl.load_workbook(str(xlsx))
    ws = wb["Sheet1"]

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=4, values_only=True):
        rows.append(row)

    if len(rows) < 2:
        print("ERROR: not enough rows in Excel"); sys.exit(1)

    # Row 2 = Tx
    tx_row = rows[1]
    tx_id = str(tx_row[0]).strip().lower() if tx_row[0] else "tx"
    tx_ex = float(tx_row[1]) if tx_row[1] else 0
    tx_ey = float(tx_row[2]) if tx_row[2] else 0
    tx_ez = float(tx_row[3]) if tx_row[3] else 0

    tx_px = tx_ex
    tx_py = tx_ez
    tx_pz = -tx_ey

    # Row 3+ = Rx positions
    rxs = []
    for row in rows[2:]:
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        if not (label.lower().startswith("rx") or label.startswith("Rx") or
                label.startswith("jie") or "接收" in label or "Rx" in label):
            continue
        try:
            ex = float(row[1]); ey = float(row[2]); ez = float(row[3])
        except (TypeError, ValueError):
            continue

        idx = extract_number(label)
        # Normalize Chinese labels to rxN
        norm_label = label
        if "接收" in label or label.startswith("jie"):
            norm_label = f"rx{idx}"
        elif label.startswith("Rx"):
            norm_label = label.lower()
        rxs.append({
            "label": norm_label, "index": idx,
            "proj_x": ex, "proj_y": ez, "proj_z": -ey
        })

    rxs.sort(key=lambda r: r["index"])

    # Build JSON
    rx_entries = []
    for r in rxs:
        rx_entries.append({
            "id": r["label"].lower(),
            "x": round(r["proj_x"], 3),
            "y": round(r["proj_y"], 3),
            "z": round(r["proj_z"], 3)
        })

    output_json = {
        "_comment": f"Auto-generated from {xlsx.name}",
        "_tx": {
            "id": tx_id,
            "x": round(tx_px, 3), "y": round(tx_py, 3), "z": round(tx_pz, 3)
        },
        "_transform": "project_x=excel_x, project_y=excel_z, project_z=-excel_y",
        "_rx_count": len(rxs),
        "rx_list": rx_entries
    }

    out_path = Path(args.output) if args.output else (xlsx.parent / f"{xlsx.stem}_rxlist.json")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(output_json, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"Tx: ({tx_px:.2f}, {tx_py:.2f}, {tx_pz:.2f})  |  {len(rxs)} Rx  |  {out_path}")


if __name__ == "__main__":
    main()
